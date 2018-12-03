from django.shortcuts import render
from .models import Datetime
from reserve.models import Student
from django.http import HttpResponse
from PresentationReserve.settings import BASE_DIR
import os
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

# Create your views here.

def sort_datatime(elem):
    return elem.datetime.year*365 + elem.datetime.month*30 + elem.datetime.day

def as_text(value):
    if value is None:
        return ""
    return str(value) + '   '

def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill

def list(request):
    # create the list
    student_list = []
    for st in Student.objects.all():
        if st.is_verified:
            student_list.append(st)
    student_list.sort(key=sort_datatime)
    book = openpyxl.Workbook()
    sheet = book.active
    cnt = 2
    sheet['A1'] = 'NUM'
    sheet['B1'] = 'EMAIL'
    sheet['C1'] = 'STUDENT NUMBER'
    sheet['D1'] = 'NAME'
    sheet['E1'] = 'TOPIC'
    sheet['F1'] = 'DATE TIME'
    num = 1
    thin = Side(border_style="thin", color="ffffff")
    double = Side(border_style="double", color="ffffff")
    border = Border(top=double, left=thin, right=thin, bottom=double)
    fill = GradientFill(stop=("000000", "000000"))
    font = Font(b=True, color="ffffff")
    al = Alignment(horizontal="center", vertical="center")
    last_datetime = ''
    for i in range(len(student_list)):
        if student_list[i].datetime != last_datetime:
            sheet['A' + str(cnt)] = student_list[i].datetime.__str__()
            style_range(sheet, 'A'+str(cnt)+':F'+str(cnt), border=border, fill=fill, font=font, alignment=al)
            cnt += 1
            num = 1
        sheet['A' + str(cnt)] = num
        sheet['B' + str(cnt)] = student_list[i].email
        sheet['C' + str(cnt)] = student_list[i].stdnum
        sheet['D' + str(cnt)] = student_list[i].name
        sheet['E' + str(cnt)] = student_list[i].topic
        sheet['F' + str(cnt)] = student_list[i].datetime.__str__()
        last_datetime = student_list[i].datetime
        num += 1
        cnt += 1
    for col in sheet.columns:
        for cell in col:
            cell.alignment = al
    for column_cells in sheet.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column].width = length

    os.chdir(os.path.join(BASE_DIR, "media"))
    if os.path.exists('list.xlsx'):
        os.remove('list.xlsx')
    book.save('list.xlsx')

    filepath = os.path.join(BASE_DIR, "media") + '/list.xlsx'
    print(filepath)
    with open(filepath, 'rb') as fp:
        data = fp.read()
    filename = 'reservation_list.xlsx'
    response = HttpResponse(content_type="application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename=%s' % filename  # force browser to download file
    response.write(data)
    return response